import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

def fetch_weather_data(api_key, cities):
   
    weather_data = []
    
    print("Fetching weather data...")
    print("-" * 60)
    
    for city in cities:
        url = f"http://api.openweathermap.org/data/2.5/weather?q={city}&appid={api_key}&units=metric&lang=en"
        
        try:
            response = requests.get(url, timeout=10)
            data = response.json()
            
            if response.status_code == 200:
                weather_data.append({
                    "City": data['name'],
                    "Weather Condition": data['weather'][0]['description'].capitalize(),
                    "Temperature (°C)": round(data['main']['temp'], 1),
                    "Max Temperature (°C)": round(data['main']['temp_max'], 1),
                    "Min Temperature (°C)": round(data['main']['temp_min'], 1),
                    "Humidity (%)": data['main']['humidity'],
                    "Wind Speed (m/s)": round(data['wind']['speed'], 1),
                    "Pressure (hPa)": data['main']['pressure']
                })
                print(f"✅ {city}: Data fetched successfully")
            else:
                print(f"❌ Failed to fetch data for {city}: {data.get('message', 'Unknown error')}")
        
        except requests.exceptions.Timeout:
            print(f"❌ Timeout error for {city}")
        except requests.exceptions.RequestException as e:
            print(f"❌ Connection error for {city}: {e}")
        except Exception as e:
            print(f"❌ Unexpected error for {city}: {e}")
    
    return weather_data


def format_excel(file_path):
    
    # Load the workbook
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define colors and styles
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    
    cell_font = Font(name='Arial', size=11)
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = thin_border
    
    # Alternating row colors for better readability
    even_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    
    # Format data rows
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column), start=2):
        for cell in row:
            cell.font = cell_font
            cell.alignment = alignment_center
            cell.border = thin_border
            
            # Apply alternating row color
            if idx % 2 == 0:
                cell.fill = even_fill
    
    # Set column widths
    column_widths = {
        'A': 18, 
        'B': 22,  
        'C': 18, 
        'D': 20, 
        'E': 20,  
        'F': 15,  
        'G': 20, 
        'H': 20   
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set row height for better readability
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 25
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Save the formatted workbook
    wb.save(file_path)
    print(f"✅ File formatted successfully!")


def get_weather_report():
    
    print("=" * 60)
    print("Weather Report Generator - Professional Edition")
    print("=" * 60)
    print()
    
    # Get API key from user
    api_key = input("Enter your OpenWeatherMap API key: ").strip()
    
    if not api_key:
        print("❌ API key is required!")
        return
    
    # Default cities
    default_cities = [ "Algiers", "Constantine", "Medea", "Paris" , "London"]
    
    # Ask user if they want to use default cities or enter custom ones
    choice = input("\nUse default cities? (y/n) [default cities: Oran, Algiers, Constantine, London, Dubai, Paris]: ").strip().lower()
    
    if choice == 'n':
        cities_input = input("Enter city names separated by commas: ").strip()
        cities = [city.strip() for city in cities_input.split(',') if city.strip()]
        
        if not cities:
            print("❌ No cities provided! Using default cities.")
            cities = default_cities
    else:
        cities = default_cities
    
    print(f"\n📍 Cities to fetch: {', '.join(cities)}")
    print()
    
    # Fetch weather data
    weather_data = fetch_weather_data(api_key, cities)
    
    if not weather_data:
        print("\n❌ No data to save.")
        return
    
    # Convert to DataFrame
    df = pd.DataFrame(weather_data)
    
    # Generate file name with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = f"Enter your desired path/weather_report_{timestamp}.xlsx"
    
    # Save to Excel
    print()
    print("-" * 60)
    print("💾 Saving data to Excel...")
    df.to_excel(file_path, index=False, engine='openpyxl')
    print(f"✅ Data saved to: {file_path}")
    
    # Format the Excel file
    print("🎨 Formatting Excel file...")
    format_excel(file_path)
    
    print()
    print("=" * 60)
    print(f"✅ Total cities processed: {len(weather_data)}")
    print(f"📁 File location: {file_path}")
    print("=" * 60)


if __name__ == "__main__":
    get_weather_report()
