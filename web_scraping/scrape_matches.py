import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def scrape_matches(date_input):

     # Website URL
    url = f"https://www.yallakora.com/match-center/?date={date_input}"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
    }

    try:
        matches_details = []
        response = requests.get(url, headers=headers)
        response.raise_for_status()

         # Find all championship sections
        soup = BeautifulSoup(response.content, "lxml")
        championships = soup.find_all("div", {"class": "matchCard"})
        
        if not championships:
            print("No matches found for this date.")
            return None

        for champ in championships:
             # Championship name
            champ_title = champ.find("h2").text.strip()
            all_matches = champ.find_all("div", {"class": "liItem"})
            
            for match in all_matches:
                team_a = match.find("div", {"class": "teamA"}).text.strip()
                team_b = match.find("div", {"class": "teamB"}).text.strip()
                score = match.find_all("span", {"class": "score"})
                score_result = f"{score[0].text.strip()} - {score[1].text.strip()}"
                match_time = match.find("span", {"class": "time"}).text.strip()

                matches_details.append({
                    "نوع البطولة": champ_title,
                    "الفريق الأول": team_a,
                    "الفريق الثاني": team_b,
                    "ميعاد المباراة": match_time,
                    "النتيجة": score_result
                })

        return matches_details

    except Exception as e:
        print(f"An error occurred: {e}")
        return None


def format_excel(file_path):
    
    # Load the workbook
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Define colors and styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    
    cell_font = Font(name='Arial', size=11)
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alignment_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
    
    # Border style
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Format header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = thin_border
    
    # Format data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = cell_font
            cell.alignment = alignment_right
            cell.border = thin_border
    
    # Set column widths
    column_widths = {
        'A': 25,  
        'B': 20,  
        'C': 20,  
        'D': 15, 
        'E': 12  
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set row height for better readability
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 25
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Save the formatted workbook
    wb.save(file_path)
    print(f"✅ File formatted successfully!")


def main():
    
    print("="*60)
    print("Yallakora Matches Scraper - Excel Edition")
    print("="*60)
    
    # Get date from user
    date_input = input("Enter date (MM/DD/YYYY): ")
    
    print(f"\nScraping matches for date: {date_input}")
    print("-"*60)
    
    # Scrape matches
    matches_details = scrape_matches(date_input)
    
    if not matches_details:
        print("❌ No data to save")
        return
    
    # Convert to DataFrame
    df = pd.DataFrame(matches_details)
    
    # Define file path
    file_path = "Enter_the_desired_path_here/matches_details.xlsx"  # Update this path as needed
  
    # Save to Excel
    df.to_excel(file_path, index=False, engine='openpyxl')
    print(f"✅ Data saved to: {file_path}")
    
    # Format the Excel file
    print("🎨 Formatting Excel file...")
    format_excel(file_path)
    
    print(f"\n✅ Total matches found: {len(matches_details)}")
    print(f"📁 File location: {file_path}")
    print("\n" + "="*60)


if __name__ == "__main__":
    main()
